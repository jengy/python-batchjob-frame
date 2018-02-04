# coding=gbk
'''
Created on 2014-5-4

@author: zhangshaojun
'''
from openpyxl.workbook import Workbook
from openpyxl.writer.excel  import ExcelWriter
from openpyxl.cell import get_column_letter 

class xlsWtUtil(object):
    '''
    classdocs
    '''
    def __init__(self, filename):
        '''
        Constructor
        '''
        self.filename = filename 
        self.wb = Workbook()
        self.ew = ExcelWriter(workbook = self.wb)
        self.wb.remove_sheet(self.wb.worksheets[0])
        return
    
    def new_sheet(self, sheetname):
        '''
        '''
        self.ws = self.wb.create_sheet()
        self.ws.title = sheetname
        self.row = 0
        return
        
    def append_rows(self, taglist, vlist):
        '''
        '''
        if self.row == 0:
            self.row = self.row + 1
            for i in range(1, len(taglist)+1):
                col = get_column_letter(i)
                self.ws.cell('%s%s'%(col, self.row)).value = '%s' % (taglist[i-1])
            
        for i in range(0, len(vlist)):
            self.row = self.row + 1
            for j in range(1, len(taglist)+1):
                if vlist[i][j-1] == None:
                    value = ''
                else:
                    value = unicode(vlist[i][j-1])
                col = get_column_letter(j)
                self.ws.cell('%s%s'%(col, self.row)).value = '%s' % value
        
        return self.row

    def save_file(self):
        '''
        '''
        self.ew.save(self.filename)
        
        return

'''
e = xlsWtUtil()
e.new_sheet('sheet-x-1')
e.append_rows(['col1'],[['value11']])

e.save_file('d:\\20140429.xlsx')
''' 